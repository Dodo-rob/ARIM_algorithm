import pandas as pd
from openpyxl import load_workbook
import numpy as np


def read_and_extract_data(file_path, usecols, start_row, end_row):
    """
    从指定的 Excel 文件中读取指定列，并从指定行范围内提取数据。

    :param file_path: Excel 文件的路径
    :param usecols: 需要读取的列的索引列表
    :param start_row: 读取数据的起始行（基于 0 索引）
    :param end_row: 读取数据的结束行（包含 end_row 行）
    :return costs: 第一列的数据（成本）
    :return prices: 第2到第8列的数据（价格）
    :return yields: 第9到第15列的数据（亩产量）
    """
    # 计算读取的行数
    nrows = end_row - start_row + 1

    # 读取指定的列和行
    data = pd.read_excel(file_path, usecols=usecols, skiprows=start_row, nrows=nrows)

    # 提取第1列作为成本
    costs = data.iloc[:, 0].values

    # 提取第2到第8列作为价格
    prices = [data.iloc[:, i].values for i in range(1, 8)]

    # 提取第9到第15列作为亩产量
    yields = [data.iloc[:, i].values for i in range(8, 15)]

    return costs, prices, yields

def delete_sheet(file_path, sheet_names):
    try:
        book = load_workbook(file_path)
        for sheet_name in sheet_names:
            if sheet_name in book.sheetnames:
                std = book[sheet_name]
                book.remove(std)
        book.save(file_path)
    except Exception as e:
        print(f"an error occurred: {e}")
        
def save_all_data_to_excel_by_year(file_path, optimal_solution, areas, area_mode, year, crop_type, prices, yields, costs):
    """
    将每一年每个作物的种植面积分配数据保存到 Excel 文件中。

    :param file_path: str, Excel 文件路径。
    :param optimal_solution: np.array, 最优解矩阵，包含每个地块每年每种作物的分配决策。
    :param areas: list, 每个地块的总面积。
    :param area_mode: int, 表示土地模式的数量（即多少个地块）。
    :param year: int, 总的种植年份数。
    :param crop_type: int, 作物种类的数量。
    :param prices: list, 各年份每种作物的价格。
    :param yields: list, 各年份每种作物的亩产量。
    :param costs: list, 每种作物的种植成本。
    """
    # 加载现有的 Excel 文件
    book = load_workbook(file_path)
    prices = np.array(prices)
    yields = np.array(yields)

    # 遍历每一年
    for k in range(year):
        # 创建一个空列表用于存储每一年的数据
        year_data = []

        for i in range(area_mode):
            # 计算每块地每种作物的面积分配
            planting_areas = np.zeros(crop_type)
            total_mianji = areas[i]
            crops_chosen = np.where(optimal_solution[i, k, :] == 1)[0]

            if len(crops_chosen) > 0:
                # 计算每种作物的利润
                crop_profits = np.zeros(len(crops_chosen))

                for idx, crop in enumerate(crops_chosen):
                    # 选择价格和亩产量
                    price = prices[0, crop] if k == 0 else \
                        prices[1, crop] if k == 1 else \
                            prices[2, crop] if k == 2 else \
                                prices[3, crop] if k == 3 else \
                                    prices[4, crop] if k == 4 else \
                                        prices[5, crop] if k == 5 else \
                                            prices[6, crop]

                    yield_per_mu = yields[0, crop] if k == 0 else \
                        yields[1, crop] if k == 1 else \
                            yields[2, crop] if k == 2 else \
                                yields[3, crop] if k == 3 else \
                                    yields[4, crop] if k == 4 else \
                                        yields[5, crop] if k == 5 else \
                                            yields[6, crop]

                    cost = costs[crop]  # 种植成本

                    # 利润计算
                    profit = (total_mianji * yield_per_mu * price) - (total_mianji * cost)
                    crop_profits[idx] = profit

                # 计算总利润
                total_profit = np.sum(crop_profits)

                # 按利润百分比分配土地面积
                if total_profit > 0:
                    profit_percentages = crop_profits / total_profit
                    for idx, crop in enumerate(crops_chosen):
                        planting_areas[crop] = profit_percentages[idx] * total_mianji

            # 添加数据到 year_data
            year_data.append([f'Plot {i + 1}'] + planting_areas.tolist())

        # 创建 DataFrame，第一行作物编号，第一列为土地类型编号
        df_year = pd.DataFrame(year_data, columns=['Plot'] + [f'Crop {j + 1}' for j in range(crop_type)])

        # 将 DataFrame 转换为列表格式，以便添加到 openpyxl 工作表
        year_data_as_list = [df_year.columns.tolist()] + df_year.values.tolist()

        # 创建新的工作表，并将数据写入其中
        sheet_name = f'Year_{2024 + k}'
        if sheet_name in book.sheetnames:
            del book[sheet_name]  # 如果工作表已存在，删除旧的工作表
        sheet = book.create_sheet(title=sheet_name)

        for row in year_data_as_list:
            sheet.append(row)  # 逐行写入 Excel 表格

    # 保存修改到文件
    book.save(file_path)
    print(f"数据成功保存到文件 {file_path}")